"""
Green Bond Market Analysis — Data Processing Script
====================================================
Cleans and aggregates Climate Bonds Initiative issuance data,
then writes the results into a multi-sheet Excel model.

Usage:
    python scripts/process_data.py --input data/cbi_newsmakers.csv

Outputs:
    Green_Bond_Market_Analysis.xlsx  (six-sheet model)
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ── colour palette ─────────────────────────────────────────────────────────
GREEN  = "1A7A4A"
LIGHT  = "E8F5E9"
HEADER = "2E7D32"

THEME_COLOURS = {
    "Green":          "1A7A4A",
    "SLB":            "FFA000",
    "Sustainability": "0288D1",
    "Social":         "C62828",
}


# ── helpers ────────────────────────────────────────────────────────────────

def header_style(ws, row: int, cols: int, title: str, colour: str = HEADER) -> None:
    """Write a styled header row."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=colour)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)


def col_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.alignment = Alignment(horizontal="center")


# ── loading ────────────────────────────────────────────────────────────────

def load(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()

    # normalise key columns
    df["Year"] = pd.to_numeric(df.get("Year", pd.Series(dtype=float)), errors="coerce")
    df["Amount (USD bn)"] = pd.to_numeric(
        df.get("Amount (USD bn)", pd.Series(dtype=float)), errors="coerce"
    )
    df["Theme"] = df.get("Theme", pd.Series(dtype=str)).str.strip()
    df["Country"] = df.get("Country", pd.Series(dtype=str)).str.strip()
    df["Sector"] = df.get("Sector", pd.Series(dtype=str)).str.strip()

    # restrict to labelled instruments and post-2015 data
    df = df[df["Year"] >= 2015].dropna(subset=["Year"])
    return df


# ── aggregations ──────────────────────────────────────────────────────────

def annual_issuance(df: pd.DataFrame) -> pd.DataFrame:
    agg = (
        df.groupby("Year")
        .agg(Deals=("Year", "count"), Volume_USD_bn=("Amount (USD bn)", "sum"))
        .reset_index()
    )
    agg["YoY_Pct"] = agg["Deals"].pct_change() * 100
    return agg.round(2)


def geographic(df: pd.DataFrame, top_n: int = 10) -> pd.DataFrame:
    return (
        df.groupby("Country")
        .agg(Deals=("Country", "count"), Volume_USD_bn=("Amount (USD bn)", "sum"))
        .sort_values("Deals", ascending=False)
        .head(top_n)
        .reset_index()
        .round(2)
    )


def sector(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.groupby("Sector")
        .agg(Deals=("Sector", "count"), Volume_USD_bn=("Amount (USD bn)", "sum"))
        .sort_values("Deals", ascending=False)
        .reset_index()
        .round(2)
    )


def theme_evolution(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.groupby(["Year", "Theme"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )


# ── sheet writers ─────────────────────────────────────────────────────────

def write_df(ws, df: pd.DataFrame, start_row: int = 2) -> None:
    """Write a DataFrame starting at start_row, with styled column headers."""
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        col_header(cell)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)


def sheet_annual(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("1. Annual Issuance")
    header_style(ws, 1, 4, "Annual Green Bond Issuance (2015–2024)")
    agg = annual_issuance(df)
    write_df(ws, agg)

    # bar chart — deal count by year
    chart = BarChart()
    chart.title = "Deal Count by Year"
    chart.y_axis.title = "Number of Deals"
    chart.x_axis.title = "Year"
    n = len(agg) + 2
    data = Reference(ws, min_col=2, min_row=2, max_row=n)
    cats = Reference(ws, min_col=1, min_row=3, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "F3")
    ws.column_dimensions["A"].width = 8
    for c in "BCDE":
        ws.column_dimensions[c].width = 18


def sheet_geo(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("2. Geography")
    header_style(ws, 1, 3, "Top 10 Countries by Deal Count")
    write_df(ws, geographic(df))
    for c in "ABC":
        ws.column_dimensions[c].width = 22


def sheet_sector(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("3. Sector Breakdown")
    header_style(ws, 1, 3, "Issuance by Sector")
    write_df(ws, sector(df))
    for c in "ABC":
        ws.column_dimensions[c].width = 25


def sheet_theme(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("4. Theme Evolution")
    header_style(ws, 1, 6, "Bond Theme Mix by Year (Deal Count)")
    write_df(ws, theme_evolution(df))
    ws.column_dimensions["A"].width = 8
    for c in "BCDEF":
        ws.column_dimensions[c].width = 16


def sheet_greenium(wb: Workbook) -> None:
    ws = wb.create_sheet("5. Greenium Analysis")
    header_style(ws, 1, 5, "Greenium — Academic Evidence Summary")

    headers = ["Study", "Year", "Market", "Greenium (bps)", "Method"]
    for col_idx, h in enumerate(headers, 1):
        col_header(ws.cell(row=2, column=col_idx, value=h))

    evidence = [
        ("Zerbib (2019)", 2019, "Developed markets", "-2 to -4", "Matched-pair OLS"),
        ("Löffler et al. (2021)", 2021, "European corporates", "-3.5", "Panel regression"),
        ("Caramichael & Rapp (2022)", 2022, "US investment grade", "-6", "Matching + event study"),
        ("Panizza et al. (2025)", 2025, "Sovereign bonds", "-5 to -8", "Synthetic control"),
        ("Banque de France (2025)", 2025, "Eurozone", "-2 to -13", "Propensity score matching"),
    ]
    for r_idx, row in enumerate(evidence, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    ws.cell(row=9, column=1, value="Matched-Pair Framework").font = Font(bold=True)
    ws.cell(row=10, column=1, value=(
        "For each green bond: identify a conventional bond from the same issuer "
        "with similar maturity (±2yr) and coupon type. Compute yield spread (conventional "
        "– green). Control for liquidity (bid-ask proxy), duration, and rating. "
        "A negative average spread = greenium exists."
    ))
    ws.merge_cells("A10:E12")
    ws.cell(row=10, column=1).alignment = Alignment(wrap_text=True)
    for c in "ABCDE":
        ws.column_dimensions[c].width = 28


def sheet_dashboard(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("0. Dashboard", 0)
    header_style(ws, 1, 4, "Green Bond Market Analysis — Summary Dashboard")

    agg = annual_issuance(df)
    total_deals = int(df.shape[0])
    countries = int(df["Country"].nunique())
    peak_year = int(agg.loc[agg["Deals"].idxmax(), "Year"])
    latest_yoy = agg["YoY_Pct"].iloc[-1]
    slb_share = df[df["Theme"] == "SLB"].shape[0] / max(total_deals, 1) * 100

    kpis = [
        ("Total Deals (2015–2024)", total_deals),
        ("Countries represented", countries),
        ("Peak year by deal count", peak_year),
        (f"YoY deal growth ({int(agg['Year'].max())})", f"{latest_yoy:.1f}%"),
        ("SLB share of all deals", f"{slb_share:.1f}%"),
    ]

    ws.cell(row=3, column=1, value="Key Metrics").font = Font(bold=True, size=11)
    for r_idx, (label, value) in enumerate(kpis, 4):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=value).font = Font(bold=True, color=GREEN)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


# ── main ──────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str) -> None:
    print(f"Loading data from {input_path} …")
    df = load(input_path)
    print(f"  {len(df):,} records loaded, {df['Year'].min():.0f}–{df['Year'].max():.0f}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_dashboard(wb, df)
    sheet_annual(wb, df)
    sheet_geo(wb, df)
    sheet_sector(wb, df)
    sheet_theme(wb, df)
    sheet_greenium(wb)

    wb.save(output_path)
    print(f"Model saved → {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Green Bond Excel model from CBI data")
    parser.add_argument("--input",  default="data/cbi_newsmakers.csv", help="Path to CBI CSV")
    parser.add_argument("--output", default="Green_Bond_Market_Analysis.xlsx", help="Output path")
    args = parser.parse_args()
    main(args.input, args.output)
